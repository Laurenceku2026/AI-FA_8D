"""Fill client 8D Excel templates for AI-FA exports."""
from __future__ import annotations

import os
import re
from datetime import datetime
from io import BytesIO
from typing import List, Optional, Tuple

from fa_template_profiles import (
    DEFAULT_8D_TEMPLATE_FILENAME,
    EIGHT_D_TEMPLATE_FILENAME,
    TEMPLATE1_8D_FILENAME,
    TEMPLATE_MODE_CUSTOM,
    TEMPLATE_MODE_DEFAULT,
    resolve_profile_template_filename,
)

try:
    from fa_template_profiles import profile_template_filename
except ImportError:  # older deployed profile module
    def profile_template_filename(mode: str, lang: str = "zh") -> str:
        names = {
            "default": ("默认-8D报告.xlsx", "Default-8D-Report.xlsx"),
            "template1": ("模板1-8D报告.xlsx", "Template-1-8D-Report.xlsx"),
        }
        zh_name, en_name = names.get(mode, ("", ""))
        return en_name if lang == "en" else zh_name


def _resolve_template_for_lang(mode: str, lang: str) -> Optional[str]:
    try:
        return resolve_profile_template_filename(mode, lang) or profile_template_filename(mode, lang)
    except TypeError:
        return resolve_profile_template_filename(mode) or profile_template_filename(mode, lang)

TEMPLATE_EXTENSIONS = (".xlsx", ".xls", ".docx")

# Example_8D.xlsx 行号（1-based，与模板严格一致）
EXAMPLE_8D_ROWS = {
    "d1": 9,
    "d2": 12,
    "d3": 15,
    "d4_why": 18,
    "d4_fishbone": 19,
    "d4_rules": 20,
    "d4_root": 21,
    "d5": 24,
    "d6": 27,
    "d7": 30,
    "d8": 33,
}
D4_COL_START = 1
D4_COL_END = 7
FISHBONE_IMAGE_COL = 3  # C 列起嵌入鱼骨图
FISHBONE_IMAGE_COL_END = 7  # G 列
FISHBONE_ROW_MIN_HEIGHT = 320.0
SIGNATURE_ROWS = (34, 35)
SIGNATURE_ROW_HEIGHT = 30.0


def _load_xlrd():
    import xlrd

    return xlrd


def _xl_copy(rb):
    from xlutils.copy import copy as xl_copy

    return xl_copy(rb)


def list_report_templates(app_key: str = "AI-FA") -> List[str]:
    here = os.path.dirname(os.path.abspath(__file__))
    folders = [
        os.path.join(here, "templates"),
        os.path.join(os.environ.get("DFSS_TEMPLATE_DIR", ""), app_key),
        os.path.join(r"C:\Users\Laurence\Technical\Project\SaaS\DFSS Report Template", app_key),
    ]
    found: List[str] = []
    for preferred in (
        profile_template_filename("default", "zh"),
        profile_template_filename("default", "en"),
        profile_template_filename("template1", "zh"),
        profile_template_filename("template1", "en"),
    ):
        if not preferred:
            continue
        try:
            resolve_template_path(preferred, app_key)
            if preferred not in found:
                found.append(preferred)
        except FileNotFoundError:
            continue
    for folder in folders:
        if not folder or not os.path.isdir(folder):
            continue
        for name in os.listdir(folder):
            if name.startswith("~$"):
                continue
            if os.path.splitext(name)[1].lower() in TEMPLATE_EXTENSIONS:
                if "8d" in name.lower() and name not in found:
                    found.append(name)
    return sorted(found)


def template_mime_type(filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".xls":
        return "application/vnd.ms-excel"
    if ext == ".docx":
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def export_report_template(
    result,
    lang: str = "zh",
    analyst_name: str = "",
    template_mode: str = TEMPLATE_MODE_DEFAULT,
    template_filename: str = "",
    template_bytes: Optional[bytes] = None,
) -> Tuple[BytesIO, str]:
    """Fill a selected 8D report template using rule mapping (no LLM)."""
    if template_mode == TEMPLATE_MODE_CUSTOM:
        if not template_bytes:
            raise ValueError(
                "请先上传自定义 8D 模板。" if lang == "zh" else "Please upload a custom 8D template first."
            )
        name = template_filename or "custom_8d.xls"
        data = fill_8d_template(
            result=result,
            lang=lang,
            analyst_name=analyst_name,
            template_bytes=template_bytes,
            template_filename=name,
        )
        return data, template_mime_type(name)

    filename = (
        template_filename
        or _resolve_template_for_lang(template_mode, lang)
        or (DEFAULT_8D_TEMPLATE_FILENAME if lang == "zh" else "Default-8D-Report.xlsx")
    )
    lower_name = filename.lower()
    if lower_name.endswith(".xls") or lower_name.endswith(".xlsx"):
        if "8d" in lower_name or "qeop" in lower_name:
            data = fill_8d_template(
                result=result,
                lang=lang,
                analyst_name=analyst_name,
                template_filename=filename,
            )
            return data, template_mime_type(filename)
    raise ValueError(
        "当前模板暂未配置自动填表规则。"
        if lang == "zh"
        else "Automatic fill rules are not configured for this template yet."
    )


def resolve_template_path(filename: str, app_key: str = "AI-FA") -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    names = [filename]
    if filename.lower().endswith(".xls"):
        names.insert(0, filename[:-4] + ".xlsx")
    elif filename.lower().endswith(".xlsx"):
        names.append(filename[:-5] + ".xls")

    candidates: List[str] = []
    for name in names:
        candidates.extend(
            [
                os.path.join(here, "templates", name),
                os.path.join(os.environ.get("DFSS_TEMPLATE_DIR", ""), app_key, name),
                os.path.join(
                    r"C:\Users\Laurence\Technical\Project\SaaS\DFSS Report Template",
                    app_key,
                    name,
                ),
            ]
        )
    for path in candidates:
        if path and os.path.isfile(path):
            return path

    # Fuzzy match for Cloud encoding / legacy names
    templates_dir = os.path.join(here, "templates")
    if os.path.isdir(templates_dir):
        markers = []
        if "默认" in filename or "default" in filename.lower():
            markers = ["默认", "8d"]
            en_markers = ["default", "8d"]
        elif "模板1" in filename or "template" in filename.lower() or "example" in filename.lower():
            markers = ["模板1", "8d"]
            en_markers = ["template-1", "8d"]
        else:
            markers = ["8d"]
            en_markers = ["8d"]
        use_en = "default-8d" in filename.lower() or "template-1-8d" in filename.lower()
        active_markers = en_markers if use_en else markers
        example_path = os.path.join(
            r"C:\Users\Laurence\Technical\Project\SaaS\DFSS Report Template",
            app_key,
            "Example_8D.xlsx",
        )
        if os.path.isfile(example_path) and ("模板1" in filename or "example" in filename.lower()):
            return example_path
        for name in os.listdir(templates_dir):
            lower = name.lower()
            if not lower.endswith((".xls", ".xlsx")):
                continue
            if active_markers and all(m.lower() in lower or m in name for m in active_markers):
                if lower.endswith(".xlsx"):
                    return os.path.join(templates_dir, name)
        for name in os.listdir(templates_dir):
            lower = name.lower()
            if not lower.endswith((".xls", ".xlsx")):
                continue
            if active_markers and all(m.lower() in lower or m in name for m in active_markers):
                return os.path.join(templates_dir, name)

    raise FileNotFoundError(f"8D template not found: {filename}")


def _clean_text(text: str) -> str:
    return re.sub(r"\*\*", "", text or "").strip()


def _join_lines(items: List[str], lang: str) -> str:
    cleaned = [_clean_text(x) for x in items if _clean_text(x)]
    if cleaned:
        body = "\n".join(f"{i + 1}. {line}" for i, line in enumerate(cleaned))
        return _format_content(body)
    return _format_content("待补充" if lang == "zh" else "Pending")


def _format_content(body: str) -> str:
    """内容单元格：先空一行再写正文（与 Example_8D 一致）。"""
    text = (body or "").strip()
    if not text:
        return "\n"
    return f"\n{text}"


def _format_titled_block(title: str, body: str) -> str:
    """D4 子节：空行 + 标题 + 空行 + 正文。"""
    text = (body or "").strip()
    if text:
        return f"\n{title}\n\n{text}"
    return f"\n{title}\n"


def _stage_name(stage: int, lang: str) -> str:
    stages_zh = {0: "正常", 1: "轻微异常", 2: "中度异常", 3: "严重故障"}
    stages_en = {0: "Normal", 1: "Minor anomaly", 2: "Moderate anomaly", 3: "Critical failure"}
    mapping = stages_zh if lang == "zh" else stages_en
    return mapping.get(stage, str(stage))


def _build_d1_team(result, analyst_name: str, lang: str) -> str:
    leader = analyst_name or ("组长" if lang == "zh" else "Team leader")
    if lang == "zh":
        lines = [
            f"1. 组长：{leader}（统筹8D流程与跨部门协调）",
            "2. 设计工程师：负责根因分析、设计变更与长期对策",
            "3. 质量工程师：负责围堵措施、验证试验与标准化",
        ]
        if result.analyst_title:
            lines[0] = f"1. 组长：{leader}（{result.analyst_title}）"
    else:
        lines = [
            f"1. Team leader: {leader} (8D coordination)",
            "2. Design engineer: root cause analysis and permanent actions",
            "3. Quality engineer: containment, verification, and standardization",
        ]
    return _format_content("\n".join(lines))


def _build_d2_problem(result, lang: str) -> str:
    symptom = _clean_text(result.symptom if lang == "zh" else result.symptom_en)
    installation = _clean_text(result.installation if lang == "zh" else result.installation_en)
    stage = _stage_name(result.failure_stage, lang)
    today = datetime.now().strftime("%Y-%m-%d")
    if lang == "zh":
        parts = [
            f"What（什么）：{symptom}",
            f"Where（何处）：{installation or '安装/使用现场'}",
            f"When（何时）：{today}",
            "Who（何人）：现场维护/质量团队",
            "Why（为何）：待根因确认",
            "How（如何发现）：运行或检验过程中发现异常",
            f"How many/How bad（程度）：{stage}（置信度 {result.root_cause_confidence:.0%}）",
        ]
    else:
        parts = [
            f"What: {symptom}",
            f"Where: {installation or 'Installation site'}",
            f"When: {today}",
            "Who: Field maintenance / quality team",
            "Why: Root cause under investigation",
            "How: Detected during operation or inspection",
            f"Severity: {stage} (confidence {result.root_cause_confidence:.0%})",
        ]
    return _format_content("\n".join(parts))


def _build_association_rules_section(result, lang: str) -> str:
    rules = getattr(result, "association_rules", None) or []
    if not rules:
        return "暂无显著关联规则，建议补充更多现场与历史数据。" if lang == "zh" else "No significant association rules; collect more field/history data."

    lines: List[str] = []
    for idx, rule in enumerate(rules[:3], 1):
        if not isinstance(rule, dict):
            continue
        antecedents = ", ".join(_clean_text(x) for x in rule.get("antecedents", []) if _clean_text(x))
        consequents = ", ".join(_clean_text(x) for x in rule.get("consequents", []) if _clean_text(x))
        explanation = _clean_text(str(rule.get("explanation", "")))
        confidence = rule.get("confidence", 0)
        try:
            conf_text = f"{float(confidence):.0%}"
        except (TypeError, ValueError):
            conf_text = str(confidence)
        if lang == "zh":
            lines.append(
                f"{idx}. 若【{antecedents or '条件待补充'}】→【{consequents or '结果待补充'}】"
                f"（置信度 {conf_text}）：{explanation}"
            )
        else:
            lines.append(
                f"{idx}. IF [{antecedents or 'TBD'}] -> [{consequents or 'TBD'}] "
                f"(confidence {conf_text}): {explanation}"
            )
    return "\n".join(lines) if lines else ("暂无显著关联规则。" if lang == "zh" else "No significant association rules.")


def _d4_title(key: str, lang: str) -> str:
    titles = {
        "why": ("【5-Why 分析】", "[5-Why Analysis]"),
        "fishbone": ("【鱼骨图分析（6M）】", "[Fishbone Analysis (6M)]"),
        "rules": ("【关联规则挖掘】", "[Association Rule Mining]"),
        "root": ("【根因结论】", "[Root Cause Conclusion]"),
    }
    zh, en = titles[key]
    return zh if lang == "zh" else en


def _build_d4_five_why_body(result, lang: str) -> str:
    if not result.five_why:
        return "待补充" if lang == "zh" else "Pending"
    lines: List[str] = []
    for item in result.five_why:
        q = _clean_text(item.question_zh if lang == "zh" else item.question_en)
        a = _clean_text(item.answer_zh if lang == "zh" else item.answer_en)
        if q or a:
            lines.append(f"Why-{item.level}: {q}\n→ {a}")
    return "\n".join(lines) if lines else ("待补充" if lang == "zh" else "Pending")


def _build_d4_fishbone_text_body(result, lang: str) -> str:
    """鱼骨图左侧 A:B 文字描述（不含标题）。"""
    cat_names_zh = {
        "Man": "人",
        "Machine": "机",
        "Material": "料",
        "Method": "法",
        "Environment": "环",
        "Measurement": "测",
    }
    lines: List[str] = []
    fishbone_dict = result.fishbone.to_dict(lang)
    for cat, causes in fishbone_dict.items():
        if not causes:
            continue
        cat_label = cat_names_zh.get(cat, cat) if lang == "zh" else cat
        cleaned = [_clean_text(c) for c in causes if _clean_text(c)]
        if cleaned:
            lines.append(f"{cat_label}：")
            for i, cause in enumerate(cleaned[:6], 1):
                lines.append(f"  {i}. {cause}")
    return "\n".join(lines) if lines else ("待补充" if lang == "zh" else "Pending")


def _build_d4_root_body(result, lang: str) -> str:
    root = _clean_text(result.root_cause_zh if lang == "zh" else result.root_cause_en)
    conf = getattr(result, "root_cause_confidence", 0) or 0
    try:
        conf_text = f"{float(conf):.0%}"
    except (TypeError, ValueError):
        conf_text = str(conf)
    if lang == "zh":
        return f"{root}\n（置信度：{conf_text}）"
    return f"{root}\n(Confidence: {conf_text})"


def _build_d4_sections(result, lang: str) -> List[str]:
    return [
        _format_titled_block(_d4_title("why", lang), _build_d4_five_why_body(result, lang)),
        _format_titled_block(_d4_title("fishbone", lang), _build_d4_fishbone_text_body(result, lang)),
        _format_titled_block(_d4_title("rules", lang), _build_association_rules_section(result, lang)),
        _format_titled_block(_d4_title("root", lang), _build_d4_root_body(result, lang)),
    ]



def _build_d6_verification(result, lang: str) -> str:
    stage = _stage_name(result.failure_stage, lang)
    if lang == "zh":
        lines = [
            "验证项目：功能测试、可靠性/耐久测试、现场回归确认",
            "验证方法：按产品标准进行型式试验，并跟踪现场批次",
            "判定准则：故障现象不再复现，关键性能指标满足规格",
            f"验证结果：功能/可靠性验证通过；失效等级：{stage}",
        ]
    else:
        lines = [
            "Items: functional test, reliability/durability test, field confirmation",
            "Method: type tests per product standard with batch follow-up",
            "Criteria: failure no longer reproduced; KPIs within specification",
            f"Result: verification passed; failure stage: {stage}",
        ]
    if result.spc_analysis:
        lines.append(
            f"SPC：{str(result.spc_analysis.get('summary', '已进行时序数据分析'))[:120]}"
            if lang == "zh"
            else f"SPC: {str(result.spc_analysis.get('summary', 'Time-series reviewed'))[:120]}"
        )
    return _format_content("\n".join(lines))


def _build_d8_recognition(lang: str) -> str:
    if lang == "zh":
        return _join_lines(
            [
                "根本原因已确认并形成闭环",
                "改进措施已定义并进入标准化",
                "经验教训已纳入知识库与流程",
            ],
            lang,
        )
    return _join_lines(
        [
            "Root cause confirmed and closed",
            "Improvement actions defined and standardized",
            "Lessons learned added to knowledge base",
        ],
        lang,
    )


def _line_display_width(text: str) -> int:
    """CJK-aware display width for Excel wrap estimation."""
    width = 0
    for ch in str(text or ""):
        width += 2 if ord(ch) > 127 else 1
    return width


def _merged_col_units(ws, col_start: int, col_end: int) -> float:
    from openpyxl.utils import get_column_letter

    total = 0.0
    for col in range(col_start, col_end + 1):
        letter = get_column_letter(col)
        total += float(ws.column_dimensions[letter].width or 8.43)
    return total


def _estimate_openpyxl_row_height(
    text: str,
    min_height: float = 15.0,
    chars_per_line: int = 52,
    ws=None,
    col_start: int = 1,
    col_end: int = 7,
) -> float:
    content = str(text or "")
    if ws is not None:
        chars_per_line = max(24, int(_merged_col_units(ws, col_start, col_end) * 0.95))
    line_count = content.count("\n") + 1
    wrapped = 0
    for line in content.splitlines() or [""]:
        width = max(1, _line_display_width(line))
        wrapped += max(1, (width + chars_per_line - 1) // chars_per_line)
    total_lines = max(line_count, wrapped)
    # Extra padding avoids Excel showing ##### when wrapped text exceeds row height.
    return min(409.0, max(min_height, 18.0 + total_lines * 15.5))


def _top_alignment():
    from openpyxl.styles import Alignment

    return Alignment(wrap_text=True, vertical="top", horizontal="left")


def _set_openpyxl_value(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = _top_alignment()


def _write_content_row(ws, row: int, text: str, chars_per_line: int = 52) -> None:
    """写入 A~G 合并内容行（不改动模板已有合并）。"""
    cell = ws.cell(row=row, column=D4_COL_START)
    cell.value = text
    cell.number_format = "General"
    cell.alignment = _top_alignment()
    min_height = float(ws.row_dimensions[row].height or 18.0)
    ws.row_dimensions[row].height = _estimate_openpyxl_row_height(
        text,
        min_height=min_height,
        chars_per_line=chars_per_line,
        ws=ws,
        col_start=D4_COL_START,
        col_end=D4_COL_END,
    )


def _column_width_pixels(ws, col_idx: int) -> int:
    from openpyxl.utils import get_column_letter

    letter = get_column_letter(col_idx)
    width = ws.column_dimensions[letter].width
    if width is None:
        width = 8.43
    return int(width * 7 + 5)


def _merged_cols_width_pixels(ws, col_start: int, col_end: int) -> int:
    return sum(_column_width_pixels(ws, col) for col in range(col_start, col_end + 1))


def _insert_fishbone_image(ws, row: int, image_bytes: Optional[bytes]) -> None:
    """在 C:G 区域嵌入鱼骨图 PNG，水平/垂直居中。"""
    ws._images = []
    if not image_bytes:
        return
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU, points_to_pixels
    except ImportError:
        return
    img = XLImage(BytesIO(image_bytes))
    max_w, max_h = 520, 400
    if img.width > max_w:
        ratio = max_w / img.width
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)
    if img.height > max_h:
        ratio = max_h / img.height
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)

    row_height_pt = float(ws.row_dimensions[row].height or FISHBONE_ROW_MIN_HEIGHT)
    row_height_emu = pixels_to_EMU(points_to_pixels(row_height_pt))
    img_w_emu = pixels_to_EMU(img.width)
    img_h_emu = pixels_to_EMU(img.height)
    row_off = max(0, (row_height_emu - img_h_emu) // 2)

    area_px = _merged_cols_width_pixels(ws, FISHBONE_IMAGE_COL, FISHBONE_IMAGE_COL_END)
    area_w_emu = pixels_to_EMU(area_px)
    col_off = max(0, (area_w_emu - img_w_emu) // 2)

    marker = AnchorMarker(
        col=FISHBONE_IMAGE_COL - 1,
        row=row - 1,
        colOff=col_off,
        rowOff=row_off,
    )
    img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(img_w_emu, img_h_emu))
    ws.add_image(img)


def _normalize_signature_rows(ws) -> None:
    """签名区各行统一行高。"""
    for row in SIGNATURE_ROWS:
        ws.row_dimensions[row].height = SIGNATURE_ROW_HEIGHT


def _write_d4_fishbone_row(ws, row: int, text: str, image_bytes: Optional[bytes]) -> None:
    """R19：A:B 文字 + C:G 鱼骨图。"""
    cell = ws.cell(row=row, column=D4_COL_START)
    cell.value = text
    cell.number_format = "General"
    cell.alignment = _top_alignment()
    text_height = _estimate_openpyxl_row_height(
        text, min_height=18.0, ws=ws, col_start=D4_COL_START, col_end=2
    )
    image_height = FISHBONE_ROW_MIN_HEIGHT if image_bytes else 0.0
    ws.row_dimensions[row].height = max(text_height, image_height, float(ws.row_dimensions[row].height or 0))
    _insert_fishbone_image(ws, row, image_bytes)


def _fill_8d_openpyxl(ws, result, lang: str, analyst_name: str) -> None:
    rows = EXAMPLE_8D_ROWS
    fishbone_image = getattr(result, "fishbone_image", None)

    product_name = _clean_text(result.product_name)
    project_name = _clean_text(result.project_name) or product_name
    today = datetime.now().strftime("%Y-%m-%d")
    stage = _stage_name(result.failure_stage, lang)

    interim = result.interim_actions_zh if lang == "zh" else result.interim_actions_en
    permanent = result.permanent_actions_zh if lang == "zh" else result.permanent_actions_en
    preventive = result.preventive_actions_zh if lang == "zh" else result.preventive_actions_en

    _set_openpyxl_value(ws, 6, 2, project_name)
    _set_openpyxl_value(ws, 6, 4, project_name)
    _set_openpyxl_value(ws, 6, 7, product_name)
    _set_openpyxl_value(ws, 7, 2, analyst_name or "-")
    _set_openpyxl_value(ws, 7, 4, stage)
    _set_openpyxl_value(ws, 7, 7, today)

    _write_content_row(ws, rows["d1"], _build_d1_team(result, analyst_name, lang))
    _write_content_row(ws, rows["d2"], _build_d2_problem(result, lang))
    _write_content_row(ws, rows["d3"], _join_lines(interim, lang))
    _write_content_row(ws, rows["d5"], _join_lines(permanent, lang))
    _write_content_row(ws, rows["d6"], _build_d6_verification(result, lang))
    _write_content_row(ws, rows["d7"], _join_lines(preventive, lang))
    _write_content_row(ws, rows["d8"], _build_d8_recognition(lang))

    d4_sections = _build_d4_sections(result, lang)
    _write_content_row(ws, rows["d4_why"], d4_sections[0])
    _write_d4_fishbone_row(ws, rows["d4_fishbone"], d4_sections[1], fishbone_image)
    _write_content_row(ws, rows["d4_rules"], d4_sections[2])
    _write_content_row(ws, rows["d4_root"], d4_sections[3])
    _normalize_signature_rows(ws)


def _fill_8d_xls_legacy(ws, rb, sh, result, lang: str, analyst_name: str) -> None:
    """Legacy .xls path for custom uploads only."""
    product_name = _clean_text(result.product_name)
    project_name = _clean_text(result.project_name) or product_name
    today = datetime.now().strftime("%Y-%m-%d")
    stage = _stage_name(result.failure_stage, lang)

    interim = result.interim_actions_zh if lang == "zh" else result.interim_actions_en
    permanent = result.permanent_actions_zh if lang == "zh" else result.permanent_actions_en
    preventive = result.preventive_actions_zh if lang == "zh" else result.preventive_actions_en

    legacy_rows = {
        "d1": (EXAMPLE_8D_ROWS["d1"] - 1, 1),
        "d2": (EXAMPLE_8D_ROWS["d2"] - 1, 1),
        "d3": (EXAMPLE_8D_ROWS["d3"] - 1, 1),
        "d5": (EXAMPLE_8D_ROWS["d5"] - 1, 1),
        "d6": (EXAMPLE_8D_ROWS["d6"] - 1, 1),
        "d7": (EXAMPLE_8D_ROWS["d7"] - 1, 1),
    }
    _write_value_preserve(ws, rb, sh, 5, 1, project_name)
    _write_value_preserve(ws, rb, sh, 5, 3, project_name)
    _write_value_preserve(ws, rb, sh, 5, 6, product_name)
    _write_value_preserve(ws, rb, sh, 6, 1, analyst_name or "-")
    _write_value_preserve(ws, rb, sh, 6, 3, stage)
    _write_value_preserve(ws, rb, sh, 6, 6, today)

    sections = {
        "d1": _build_d1_team(result, analyst_name, lang),
        "d2": _build_d2_problem(result, lang),
        "d3": _join_lines(interim, lang),
        "d5": _join_lines(permanent, lang),
        "d6": _build_d6_verification(result, lang),
        "d7": _join_lines(preventive, lang),
    }
    for key, content in sections.items():
        row, row_span = legacy_rows[key]
        _write_section_cell(ws, rb, sh, row, row_span, content, force_top=True)
    d4_rows = (
        EXAMPLE_8D_ROWS["d4_why"] - 1,
        EXAMPLE_8D_ROWS["d4_fishbone"] - 1,
        EXAMPLE_8D_ROWS["d4_rules"] - 1,
        EXAMPLE_8D_ROWS["d4_root"] - 1,
    )
    for row, content in zip(d4_rows, _build_d4_sections(result, lang)):
        _write_section_cell(ws, rb, sh, row, 1, content, force_top=True)
    _write_section_cell(
        ws, rb, sh, EXAMPLE_8D_ROWS["d8"] - 1, 1, _build_d8_recognition(lang), force_top=True
    )


def _xlwt_border_line(style: int) -> int:
    return {0: 0, 1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6}.get(style, 1)


def _style_from_template_xf(rb, xf_index: int, force_top: bool = True):
    """Copy template xf (borders/alignment) into an xlwt style."""
    import xlwt

    xf = rb.xf_list[xf_index]
    style = xlwt.XFStyle()

    align = xlwt.Alignment()
    horz_map = {0: xlwt.Alignment.HORZ_GENERAL, 1: xlwt.Alignment.HORZ_LEFT, 2: xlwt.Alignment.HORZ_CENTER, 3: xlwt.Alignment.HORZ_RIGHT}
    align.horz = horz_map.get(xf.alignment.hor_align, xlwt.Alignment.HORZ_LEFT)
    if force_top:
        align.vert = xlwt.Alignment.VERT_TOP
    else:
        vert_map = {0: xlwt.Alignment.VERT_TOP, 1: xlwt.Alignment.VERT_CENTER, 2: xlwt.Alignment.VERT_BOTTOM}
        align.vert = vert_map.get(xf.alignment.vert_align, xlwt.Alignment.VERT_TOP)
    align.wrap = (
        xlwt.Alignment.WRAP_AT_RIGHT
        if xf.alignment.text_wrapped
        else xlwt.Alignment.NOT_WRAP_AT_RIGHT
    )
    style.alignment = align

    borders = xlwt.Borders()
    b = xf.border
    borders.left = _xlwt_border_line(b.left_line_style)
    borders.right = _xlwt_border_line(b.right_line_style)
    borders.top = _xlwt_border_line(b.top_line_style)
    borders.bottom = _xlwt_border_line(b.bottom_line_style)
    style.borders = borders
    return style


def _write_value_preserve(ws, rb, sh, row: int, col: int, value: str, force_top: bool = True) -> None:
    xf_index = sh.cell(row, col).xf_index
    style = _style_from_template_xf(rb, xf_index, force_top=force_top)
    ws.write(row, col, value, style)


def _write_section_cell(
    ws,
    rb,
    sh,
    row: int,
    row_span: int,
    text: str,
    force_top: bool = True,
) -> None:
    """Legacy .xls: write top-left only (never write_merge) to avoid breaking borders."""
    col_start = 0
    xf_index = sh.cell(row, col_start).xf_index
    style = _style_from_template_xf(rb, xf_index, force_top=force_top)
    ws.write(row, col_start, text, style)

    min_height = sum(
        (sh.rowinfo_map[row + i].height if row + i in sh.rowinfo_map else 400)
        for i in range(max(row_span, 1))
    )
    height = max(min_height, _estimate_row_height(text))
    row_obj = ws.row(row)
    row_obj.height = min(20000, int(height))
    row_obj.height_mismatch = True


def _estimate_row_height(text: str, base: int = 480, per_line: int = 280, chars_per_line: int = 42) -> int:
    content = str(text or "")
    line_count = content.count("\n") + 1
    wrapped = sum(max(1, (len(line) + chars_per_line - 1) // chars_per_line) for line in content.splitlines() or [""])
    total_lines = max(line_count, wrapped)
    return min(12000, base + total_lines * per_line)


def fill_8d_template(
    result,
    lang: str = "zh",
    analyst_name: str = "",
    template_filename: str = DEFAULT_8D_TEMPLATE_FILENAME,
    template_bytes: Optional[bytes] = None,
) -> BytesIO:
    """Fill the client 8D template and return an in-memory workbook."""
    if template_bytes:
        source = BytesIO(template_bytes)
        name = template_filename.lower()
    else:
        source = resolve_template_path(template_filename, "AI-FA")
        name = source.lower()

    if name.endswith(".xlsx") or (template_bytes and template_filename.lower().endswith(".xlsx")):
        from openpyxl import load_workbook

        wb = load_workbook(source)
        ws = wb.active
        _fill_8d_openpyxl(ws, result, lang, analyst_name)
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    xlrd = _load_xlrd()
    if template_bytes:
        rb = xlrd.open_workbook(file_contents=template_bytes, formatting_info=True)
    else:
        rb = xlrd.open_workbook(source, formatting_info=True)
    wb = _xl_copy(rb)
    ws = wb.get_sheet(0)
    sh = rb.sheet_by_index(0)
    _fill_8d_xls_legacy(ws, rb, sh, result, lang, analyst_name)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
