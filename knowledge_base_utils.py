"""Shared Supabase knowledge base utilities for AI-DQA and AI-FA."""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional
from urllib.parse import quote

import pandas as pd
import requests

KNOWLEDGE_CATEGORIES = ["光学", "机械", "材料", "热学", "电气", "控制", "其他"]
KB_CATEGORY_HEADERS = [
    "光学 / Optical",
    "机械 / Mechanical",
    "材料 / Material",
    "热学 / Thermal",
    "电气 / Electrical",
    "控制 / Control",
    "其他 / Other",
]
KB_HEADER_ROW = 3
KB_DATA_START_ROW = 4

_CATEGORY_ALIASES = {
    "光学 / Optical": "光学",
    "机械 / Mechanical": "机械",
    "材料 / Material": "材料",
    "热学 / Thermal": "热学",
    "电气 / Electrical": "电气",
    "控制 / Control": "控制",
    "其他 / Other": "其他",
    "光学": "光学",
    "机械": "机械",
    "材料": "材料",
    "热学": "热学",
    "电气": "电气",
    "控制": "控制",
    "其他": "其他",
    "Other": "其他",
    "other": "其他",
}


def is_chinese(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", text or ""))


def normalize_category(label: Any) -> Optional[str]:
    if label is None:
        return None
    text = str(label).strip()
    if not text or text.lower() == "nan":
        return None
    if text in _CATEGORY_ALIASES:
        return _CATEGORY_ALIASES[text]
    if "/" in text:
        zh_part = text.split("/", 1)[0].strip()
        if zh_part in KNOWLEDGE_CATEGORIES:
            return zh_part
        en_part = text.split("/", 1)[1].strip().lower()
        if en_part == "other":
            return "其他"
    return None


def parse_wide_kb_excel(file_bytes: bytes) -> List[Dict[str, str]]:
    """Parse enterprise wide-format workbook: headers on row 3, data from row 4."""
    from openpyxl import load_workbook

    worksheet = load_workbook(io.BytesIO(file_bytes)).active
    categories_by_col: Dict[int, str] = {}
    for col_idx in range(1, len(KNOWLEDGE_CATEGORIES) + 1):
        category = normalize_category(worksheet.cell(KB_HEADER_ROW, col_idx).value)
        if category:
            categories_by_col[col_idx] = category

    if "其他" not in categories_by_col.values():
        return []

    rows: List[Dict[str, str]] = []
    for row_idx in range(KB_DATA_START_ROW, worksheet.max_row + 1):
        for col_idx, category in categories_by_col.items():
            raw_value = worksheet.cell(row_idx, col_idx).value
            if raw_value is None:
                continue
            content = str(raw_value).strip()
            if not content or content.lower() == "nan":
                continue
            rows.append({"category": category, "content": content})
    return rows


class SupabaseKnowledgeDB:
    """Supabase knowledge base with platform + tenant scopes and bilingual search."""

    def __init__(
        self,
        supabase_url: str,
        service_role_key: str,
        translate_to_en: Callable[[str], str],
        translate_to_zh: Callable[[str], str],
        ui_lang_getter: Callable[[], str] = lambda: "zh",
        *,
        organization_id: Optional[str] = None,
        kb_scope: str = "platform",
        include_tenant_kb: bool = True,
    ):
        self.supabase_url = supabase_url
        self.service_role_key = service_role_key
        self._translate_to_en = translate_to_en
        self._translate_to_zh = translate_to_zh
        self._ui_lang_getter = ui_lang_getter
        self.organization_id = organization_id
        self.kb_scope = kb_scope
        self.include_tenant_kb = include_tenant_kb
        self.categories = list(KNOWLEDGE_CATEGORIES)
        self.headers = {
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Content-Type": "application/json",
        }
        self.knowledge_zh: Dict[str, List[str]] = {cat: [] for cat in self.categories}
        self.knowledge_en: Dict[str, List[str]] = {cat: [] for cat in self.categories}
        self._load_cache()

    def _fetch_scope_rows(self, scope: str, organization_id: Optional[str] = None) -> List[Dict[str, Any]]:
        if not self.supabase_url or not self.service_role_key:
            return []
        if scope == "tenant" and not organization_id:
            return []

        query = f"select=id,category,content,content_en,scope,organization_id&scope=eq.{quote(scope)}&order=id"
        if scope == "tenant" and organization_id:
            query += f"&organization_id=eq.{quote(str(organization_id), safe='')}"

        try:
            response = requests.get(
                f"{self.supabase_url}/rest/v1/knowledge_base?{query}",
                headers=self.headers,
                timeout=15,
            )
            if response.status_code == 200:
                return response.json()
        except Exception as exc:
            print(f"加载知识库失败({scope}): {exc}")
        return []

    def _load_cache(self) -> None:
        self.knowledge_zh = {cat: [] for cat in self.categories}
        self.knowledge_en = {cat: [] for cat in self.categories}
        if not self.supabase_url or not self.service_role_key:
            return

        rows: List[Dict[str, Any]] = []
        rows.extend(self._fetch_scope_rows("platform"))
        if self.include_tenant_kb and self.organization_id:
            rows.extend(self._fetch_scope_rows("tenant", self.organization_id))

        for row in rows:
            cat = normalize_category(row.get("category")) or row.get("category")
            if cat not in self.knowledge_zh:
                continue
            if row.get("content"):
                self.knowledge_zh[cat].append(row.get("content"))
            if row.get("content_en"):
                self.knowledge_en[cat].append(row.get("content_en"))

    def get_knowledge(self, category: str, lang: str = "zh") -> List[str]:
        if lang == "zh":
            return self.knowledge_zh.get(category, [])
        return self.knowledge_en.get(category, [])

    def get_knowledge_by_category(self, category: str) -> List[str]:
        return self.get_knowledge(category, self._ui_lang_getter())

    def search_knowledge_dual(self, query: str, lang: str) -> List[str]:
        if not query or not query.strip():
            return []

        results: List[str] = []
        query_lower = query.lower()

        for cat in self.categories:
            for item in self.get_knowledge(cat, lang):
                if query_lower in item.lower() and item not in results:
                    results.append(item)

        other_lang = "en" if lang == "zh" else "zh"
        trans_query = (
            self._translate_to_en(query) if lang == "zh" else self._translate_to_zh(query)
        )
        trans_lower = (trans_query or "").lower()
        if trans_lower:
            for cat in self.categories:
                for item in self.get_knowledge(cat, other_lang):
                    if trans_lower in item.lower() and item not in results:
                        results.append(item)

        return results[:10]

    def search_knowledge(self, keywords: str, limit: int = 10) -> List[str]:
        query_lang = "zh" if is_chinese(keywords) else "en"
        return self.search_knowledge_dual(keywords, query_lang)[:limit]

    def search_knowledge_full(self, query: str, limit: int = 10) -> List[str]:
        if not query or not query.strip():
            return []

        primary_lang = "zh" if is_chinese(query) else "en"
        other_lang = "en" if primary_lang == "zh" else "zh"
        translated = (
            self._translate_to_en(query) if primary_lang == "zh" else self._translate_to_zh(query)
        )

        merged: List[str] = []
        for item in self.search_knowledge_dual(query, primary_lang):
            if item not in merged:
                merged.append(item)
        for item in self.search_knowledge_dual(translated, other_lang):
            if item not in merged:
                merged.append(item)
        return merged[:limit]

    def _bilingualize(self, content: str) -> tuple[str, str]:
        if is_chinese(content):
            return content, self._translate_to_en(content)
        return self._translate_to_zh(content), content

    def add_knowledge(self, category: str, content: str) -> bool:
        category = normalize_category(category) or category
        if category not in self.categories:
            return False

        zh_text, en_text = self._bilingualize(content)
        payload: Dict[str, Any] = {
            "category": category,
            "content": zh_text,
            "content_en": en_text,
            "scope": self.kb_scope,
            "created_at": datetime.now().isoformat(),
        }
        if self.kb_scope == "tenant" and self.organization_id:
            payload["organization_id"] = self.organization_id

        try:
            response = requests.post(
                f"{self.supabase_url}/rest/v1/knowledge_base",
                headers=self.headers,
                json=payload,
                timeout=15,
            )
            if response.status_code in (200, 201, 204):
                self._load_cache()
                return True
        except Exception as exc:
            print(f"添加知识库失败: {exc}")
        return False

    def delete_knowledge(self, category: str, content: str) -> bool:
        lang = self._ui_lang_getter()
        category = normalize_category(category) or category
        try:
            scope_filter = f"&scope=eq.{self.kb_scope}"
            if self.kb_scope == "tenant" and self.organization_id:
                scope_filter += f"&organization_id=eq.{quote(str(self.organization_id), safe='')}"
            if lang == "zh":
                url = (
                    f"{self.supabase_url}/rest/v1/knowledge_base"
                    f"?category=eq.{quote(category)}&content=eq.{quote(content)}{scope_filter}"
                )
            else:
                url = (
                    f"{self.supabase_url}/rest/v1/knowledge_base"
                    f"?category=eq.{quote(category)}&content_en=eq.{quote(content)}{scope_filter}"
                )
            response = requests.get(url, headers=self.headers, timeout=15)
            if response.status_code == 200 and response.json():
                record_id = response.json()[0]["id"]
                delete_resp = requests.delete(
                    f"{self.supabase_url}/rest/v1/knowledge_base?id=eq.{record_id}",
                    headers=self.headers,
                    timeout=15,
                )
                if delete_resp.status_code in (200, 204):
                    self._load_cache()
                    return True
        except Exception as exc:
            print(f"删除知识库失败: {exc}")
        return False

    def clear_category(self, category: str) -> bool:
        category = normalize_category(category) or category
        try:
            query = f"category=eq.{quote(category)}&scope=eq.{self.kb_scope}"
            if self.kb_scope == "tenant" and self.organization_id:
                query += f"&organization_id=eq.{quote(str(self.organization_id), safe='')}"
            response = requests.delete(
                f"{self.supabase_url}/rest/v1/knowledge_base?{query}",
                headers=self.headers,
                timeout=15,
            )
            if response.status_code in (200, 204):
                self._load_cache()
                return True
        except Exception as exc:
            print(f"清空分类失败: {exc}")
        return False

    def clear_knowledge_category(self, category: str) -> bool:
        return self.clear_category(category)

    def clear_scope(self) -> bool:
        try:
            query = f"scope=eq.{self.kb_scope}"
            if self.kb_scope == "tenant" and self.organization_id:
                query += f"&organization_id=eq.{quote(str(self.organization_id), safe='')}"
            response = requests.delete(
                f"{self.supabase_url}/rest/v1/knowledge_base?{query}",
                headers=self.headers,
                timeout=15,
            )
            if response.status_code in (200, 204):
                self._load_cache()
                return True
        except Exception as exc:
            print(f"清空知识库失败: {exc}")
        return False

    def import_rows(self, rows: List[Dict[str, str]], *, replace_existing: bool = True) -> int:
        if replace_existing:
            self.clear_scope()
        imported = 0
        for row in rows:
            if self.add_knowledge(row["category"], row["content"]):
                imported += 1
        self._load_cache()
        return imported

    def get_all_knowledge(self) -> Dict[str, List[str]]:
        if self._ui_lang_getter() == "zh":
            return self.knowledge_zh
        return self.knowledge_en

    def export_to_dataframe(self) -> pd.DataFrame:
        max_len = max((len(self.knowledge_zh.get(cat, [])) for cat in self.categories), default=0)
        export_data = {}
        for cat in self.categories:
            items = self.knowledge_zh.get(cat, [])
            export_data[cat] = items + [""] * (max_len - len(items))
        return pd.DataFrame(export_data)

    def import_from_dataframe(self, df: pd.DataFrame) -> int:
        rows: List[Dict[str, str]] = []
        for column in df.columns:
            category = normalize_category(column)
            if not category:
                continue
            for item in df[column].dropna():
                text = str(item).strip()
                if text:
                    rows.append({"category": category, "content": text})
        return self.import_rows(rows, replace_existing=True)

    def import_from_excel_bytes(self, file_bytes: bytes) -> int:
        rows = parse_wide_kb_excel(file_bytes)
        if not rows:
            legacy_df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
            return self.import_from_dataframe(legacy_df)
        return self.import_rows(rows, replace_existing=True)
