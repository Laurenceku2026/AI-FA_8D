"""Build English 8D template shells from Chinese templates (same layout)."""
from __future__ import annotations

import os
import re
from copy import copy

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATES = os.path.join(HERE, "templates")

FA_REPLACEMENTS = [
    ("8D报告", "8D Report"),
    ("8D表格", "8D Report"),
    ("客户名称(CUSTOMER NAME）：", "Customer Name:"),
    ("客户名称(CUSTOMER NAME）:", "Customer Name:"),
    ("部件编号（PART NUMBER）:", "Part Number:"),
    ("部件名称（PART DESCRIPTION）：", "Part Description:"),
    ("客户代码（CUSTOMER CODE）", "Customer Code"),
    ("失效率（FAILURE\u00a0RATE）", "Failure Rate"),
    ("失效率（FAILURE RATE）", "Failure Rate"),
    ("日期（DATE）：", "Date:"),
    ("1.改善小组成员（IMPROVEMEN TEAM MEMBERS）", "1. Improvement Team Members"),
    ("签名：", "Signature:"),
    ("时间：", "Date:"),
    ("2.  问题描述（DESCRIBE THE PROBLEM）:", "2. Describe the Problem:"),
    ("3短期对策(Short\u00a0Term\u00a0Corrective\u00a0Action)", "3. Short-Term Corrective Action"),
    ("3短期对策(Short Term Corrective Action)", "3. Short-Term Corrective Action"),
    ("4.原因分析（Cause\u00a0Analysis）：", "4. Cause Analysis:"),
    ("4.原因分析（Cause Analysis）：", "4. Cause Analysis:"),
    ("5.长期对策 （ LONG-TERM PERMANENT CORRECTIVE ACTION）：", "5. Long-Term Permanent Corrective Action:"),
    ("6.对策验证（Verification\u00a0of\u00a0Corrective\u00a0Actions）：", "6. Verification of Corrective Actions:"),
    ("6.对策验证（Verification of Corrective Actions）：", "6. Verification of Corrective Actions:"),
    ("7.形成标准化（Standardization）", "7. Standardization"),
    ("8.祝贺小组(Congratulate\u00a0Team)", "8. Congratulate Team"),
    ("8.祝贺小组(Congratulate Team)", "8. Congratulate Team"),
    ("审批：", "Approval:"),
    ("表单编号：模板-1   保存年限：2年   版本：A1", "Form No.: Template-1   Retention: 2 years   Rev.: A1"),
    ("表单编号：默认   保存年限：2年   版本：A1", "Form No.: Default   Retention: 2 years   Rev.: A1"),
]


def _translate(value: str, rules: list[tuple[str, str]]) -> str:
    text = str(value)
    for src, dst in rules:
        text = text.replace(src, dst)
    return text


def _translate_workbook(src_path: str, dst_path: str, rules: list[tuple[str, str]]) -> None:
    wb = load_workbook(src_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                translated = _translate(cell.value, rules)
                if translated != cell.value:
                    cell.value = translated
        if ws.title in ("8D表格",):
            ws.title = "8D Report"
    wb.save(dst_path)
    print("Wrote", dst_path)


def main() -> None:
    pairs = [
        ("默认-8D报告.xlsx", "Default-8D-Report.xlsx"),
        ("模板1-8D报告.xlsx", "Template-1-8D-Report.xlsx"),
    ]
    for src_name, dst_name in pairs:
        src = os.path.join(TEMPLATES, src_name)
        dst = os.path.join(TEMPLATES, dst_name)
        if not os.path.isfile(src):
            raise FileNotFoundError(src)
        _translate_workbook(src, dst, FA_REPLACEMENTS)


if __name__ == "__main__":
    main()
